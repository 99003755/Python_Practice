# -*- coding: utf-8 -*-
"""
Created on Tue Mar 16 06:25:50 2021

@author: Rohan Roy
"""
from openpyxl import load_workbook
from openpyxl.styles import Font

path = "D:\Python_Practice\LnT.xlsx"
wb = load_workbook(path)
found = 0
name = input("Enter Name: ")
PS = eval(input("Enter PS Number: "))
email = input(" Enter email: ")

data = []
head = []
#def sheet(sh_name ,name, PS, email):
for sheet in wb.sheetnames:
    print("IN SHEET")

# SHEET 1 Phone Number
# ws=wb.get_sheet_by_name('Sheet1')
    ws = wb[sheet]
    s = ws.max_row  # variable to store max rows for sl num
    col = ws.max_column
    #maxr = ws.max_row
    for i in range(1, s + 1):
        if ws.cell(row=i, column=1).value == name and ws.cell(row=i, column=2).value == PS \
                and ws.cell(row=i, column=3).value == email:
            if sheet == "Sheet0":
                break
            if sheet != 'Sheet1':

                for k in range(4,col+1):
                    data.append(ws.cell(row=i, column=k).value)
                    head.append(ws.cell(row=1, column=k).value)
            else:
                for j in range(1, col+1):
                    #print(ws.cell(row=i, column=j).value)
                    data.append(ws.cell(row=i, column=j).value)
                    head.append(ws.cell(row=1, column=j).value)
            found = 1

    print(data)


# Master Sheet (Sheet0)

if found == 1:
    if 'Sheet0' not in wb.sheetnames:
        # head = []
        #head = ['Name', 'PS Number', 'Email', 'Phone Number', 'Batch', 'Location', 'BU', 'XYZ']
        ws = wb.create_sheet('Sheet0')
        print("CREATING")
        s = ws.max_row  # variable to store max rows for sl num
        for i in range(1, len(head)+1):
            ws.cell(row=1, column=i).value = head[i - 1]
        for i in range(1, len(head)+1):
            clr = ws.cell(row=1, column=i)
            clr.font = Font(bold=True)
        for i in range(1, len(head)+1):
            ws.cell(row=s + 1, column=i).value = data[i - 1]
        wb.save(path)
    else:
        # ws = wb.get_sheet_by_name('Sheet0')
        ws = wb['Sheet0']
        s = ws.max_row
        for i in range(1, len(head)+1):
            ws.cell(row=s + 1, column=i).value = data[i - 1]
        wb.save(path)
if found == 0:
    print("DATA NOT FOUND")



from openpyxl.chart import BarChart, Series, Reference
wb = load_workbook(path)
ws = wb['Sheet0']

chart1 = BarChart()
#chart1.type = "col"
#chart1.style = 11
chart1.title = "EXCEL DATA"
chart1.y_axis.title = 'Attributes'
chart1.x_axis.title = 'X-axis'
print(col)
data = Reference(ws, min_col=3, min_row=2, max_row=10, max_col=4)

chart1.add_data(data, titles_from_data=True)
#chart1.set_categories(cats)

chart1.shape = 4
ws.add_chart(chart1, "Z3")
wb.save(path)

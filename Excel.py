# -*- coding: utf-8 -*-
"""
Created on Tue Mar 16 06:25:50 2021

@author: Rohan Roy
"""
from openpyxl import load_workbook
from openpyxl.styles import Font

path = "D:\Python_Practice\LnT.xlsx"
wb = load_workbook(path)

name = input("Enter Name: ")
PS = eval(input("Enter PS Number: "))
email = input(" Enter email: ")

# SHEET 1 Phone Number
# ws=wb.get_sheet_by_name('Sheet1')
ws = wb['Sheet1']
s = ws.max_row  # variable to store max rows for sl num
maxr = ws.max_row
data = []

for i in range(1, s + 1):
    if ws.cell(row=i, column=1).value == name and ws.cell(row=i, column=2).value == PS \
            and ws.cell(row=i, column=3).value == email:
        for j in range(1, 5):
            print(ws.cell(row=i, column=j).value)
            data.append(ws.cell(row=i, column=j).value)
print(data)

# SHEET 2 Batch
# ws=wb.get_sheet_by_name('Sheet2')
ws = wb['Sheet2']
s = ws.max_row  # variable to store max rows for sl num
maxr = ws.max_row
for i in range(1, s + 1):
    if ws.cell(row=i, column=1).value == name and ws.cell(row=i, column=2).value == PS \
            and ws.cell(row=i, column=3).value == email:
        print(ws.cell(row=i, column=4).value)
        data.append(ws.cell(row=i, column=4).value)
print(data)

# SHEET 3 Location
# ws=wb.get_sheet_by_name('Sheet2')
ws = wb['Sheet3']
s = ws.max_row  # variable to store max rows for sl num
maxr = ws.max_row
for i in range(1, s + 1):
    if ws.cell(row=i, column=1).value == name and ws.cell(row=i, column=2).value == PS \
            and ws.cell(row=i, column=3).value == email:
        print(ws.cell(row=i, column=4).value)
        data.append(ws.cell(row=i, column=4).value)
print(data)

# SHEET 4 BU
# ws=wb.get_sheet_by_name('Sheet2')
ws = wb['Sheet4']
s = ws.max_row  # variable to store max rows for sl num
maxr = ws.max_row
for i in range(1, s + 1):
    if ws.cell(row=i, column=1).value == name and ws.cell(row=i, column=2).value == PS \
            and ws.cell(row=i, column=3).value == email:
        print(ws.cell(row=i, column=4).value)
        data.append(ws.cell(row=i, column=4).value)
print(data)

# SHEET 5 XYZ
# ws=wb.get_sheet_by_name('Sheet2')
ws = wb['Sheet5']
s = ws.max_row  # variable to store max rows for sl num
maxr = ws.max_row
for i in range(1, s + 1):
    if ws.cell(row=i, column=1).value == name and ws.cell(row=i, column=2).value == PS \
            and ws.cell(row=i, column=3).value == email:
        print(ws.cell(row=i, column=4).value)
        data.append(ws.cell(row=i, column=4).value)
print(data)

# Master Sheet (Sheet0)
if 'Sheet0' not in wb.sheetnames:
    # head = []
    head = ['Name', 'PS Number', 'Email', 'Phone Number', 'Batch', 'Location', 'BU', 'XYZ']
    ws = wb.create_sheet('Sheet0')
    print("CREATING")
    s = ws.max_row  # variable to store max rows for sl num
    for i in range(1, 9):
        ws.cell(row=1, column=i).value = head[i - 1]
    for i in range(1, 9):
        clr = ws.cell(row=1, column=i)
        clr.font = Font(bold=True)
    for i in range(1, 9):
        ws.cell(row=s + 1, column=i).value = data[i - 1]
    wb.save(path)

else:
    # ws = wb.get_sheet_by_name('Sheet0')
    ws = wb['Sheet0']
    s = ws.max_row
    for i in range(1, 9):
        ws.cell(row=s + 1, column=i).value = data[i - 1]
    wb.save(path)
